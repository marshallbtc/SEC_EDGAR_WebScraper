# import libraries
import re, requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# go ahead and make an excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Analysis"

# TODO: Short term goal is to be able to plug in the company ticker without manually entering
# the CIK or Accession number.
CIK = "1318605"
ACC_num = "0001564590-20-004475"

"""
Grab the Filing XML Summary:
"""
def get_filing_summary():
    # define the base url needed to create the file url.
    base_url = r"https://www.sec.gov/"

    # convert a normal url to a 10k document landing page url
    normal_url = base_url + "Archives/edgar/data/" + CIK + "/" + ACC_num + ".txt"
    DocLandPage_url = normal_url.replace('-','').replace('.txt','/index.json')
    #DocLandPage_url = r"https://www.sec.gov/Archives/edgar/data/1265107/000126510719000004/index.json"

    # request the url and decode it.
    content = requests.get(DocLandPage_url).json()

    for file in content['directory']['item']:
        
        # Grab the filing summary and create a new url leading to the file so we can download it.
        if file['name'] == 'FilingSummary.xml':

            xml_summary = base_url + content['directory']['name'] + "/" + file['name']
            
            print('-' * 100)
            print('File Name: ' + file['name'])
            print('File Path: ' + xml_summary)

    return xml_summary

"""
Parse the Filing Summary:
"""
def parse_filing_summary(xml_summary):
    # define a new base url that represents the filing folder. This will come
    # in handy when we need to download the reports.
    base_url = xml_summary.replace('FilingSummary.xml', '')

    # request and parse the content
    content = requests.get(xml_summary).content
    soup = BeautifulSoup(content, 'lxml')

    # find the 'myreports' tag because this contains all the individual reports submitted.
    reports = soup.find('myreports')

    # I want a list to store all the individual components of the report, so create the master list.
    master_reports = []

    # loop through each report in the 'myreports' tag but avoid the last one as this will cause an error.
    for report in reports.find_all('report')[:-1]:

        # let's create a dictionary to store all the different parts we need.
        report_dict = {}
        report_dict['name_short'] = report.shortname.text
        report_dict['name_long'] = report.longname.text
        report_dict['position'] = report.position.text
        report_dict['category'] = report.menucategory.text
        report_dict['url'] = base_url + report.htmlfilename.text

        # append the dictionary to the master list.
        master_reports.append(report_dict)

        '''# print the info to the user.
        print('-'*100)
        print("BASE URL:\n",base_url + report.htmlfilename.text)
        #print("LONGNAME:\n",report.longname.text)
        print("SHORTNAME:\n",report.shortname.text)
        #print("MENU CATEGORY:\n",report.menucategory.text)
        print("POSITION:\n",report.position.text)'''

    return master_reports
    
"""
Grab the Financial Statements:
"""
def grab_financial_statements(master_reports):
    # create the list to hold the statement urls
    statements_url = []
    print("\nmaster_reports:\n", master_reports)

    # define the statements we want to look for (defined as item1 - item4)
    for v in master_reports:
        # BALANCE SHEET
        if "Balance" in v['name_short']:
            item1 = r"" + v['name_short']
            break
        #else:
        #    print("BALANCE SHEET NAME ERROR: \nNeed to alter the above if statement's string according to the company's Balance Sheet name_short\n")
        #    item1 = "ERROR"
        #    break
    for v in master_reports:
        # INCOME STATEMENT
        if "Operations" in v['name_short']:
            item2 = r"" + v['name_short']
            break
        elif "Income" in v['name_short']:
            item2 = r"" + v['name_short']
            break
        #else:
        #    print("INCOME STATEMENT NAME ERROR: \nNeed to alter the above if statement's string according to the company's Income Statement name_short\n")
        #    item2 = "ERROR"
        #    break
    for v in master_reports:
        # STATEMENT OF CASH FLOWS
        if "Cash" in v['name_short']:
            item3 = r"" + v['name_short']
            break
        #else:
        #    print("STATEMENT OF CASH FLOWS NAME ERROR: \nNeed to alter the above if statement's string according to the company's Statement of Cash Flows name_short\n")
        #    item3 = "ERROR"
        #    break
    for v in master_reports:
        # STATEMENT OF STOCKHOLDERS EQUITY
        if "Equity" in v['name_short']:
            item4 = r"" + v['name_short']
            break
        #else:
        #    print("STATEMENT OF STOCKHOLDERS EQUITY NAME ERROR: \nNeed to alter the above if statement's string according to the company's Statement of Stockholders' Equity name_short\n")
        #    item4 = "ERROR"
        #    break
    print("item1: ",item1)
    print("item2: ",item2)
    print("item3: ",item3)
    print("item4: ",item4)

    # store the statement names in a list
    report_list = [item1, item2, item3, item4]

    for reportDict in master_reports:

        # if the short name can be found in the report list.
        if reportDict['name_short'] in report_list:

            # print some info and store it in the statements url.
            print('-'*100)
            print(reportDict['name_short'])
            print(reportDict['url'])
                
            statements_url.append(reportDict['url'])
        
    #print("Balance sheet name:\n",report_list[0])
    return statements_url

"""
Scrape the Financial Statements:
"""
def scrape_financial_statements(statements_url):
    # let's assume we want all the statements in a single data set.
    statements_data = []

    # loop through each statement url
    for statement in statements_url:

        # define a dictionary that will store the different parts of the statement.
        statement_data = {}
        statement_data['headers'] = []
        statement_data['sections'] = []
        statement_data['data'] = []
        
        # request the statement file content
        content = requests.get(statement).content
        report_soup = BeautifulSoup(content, 'html')

        # find all the rows, figure out what type of row it is, parse the elements, and store in the statement file list.
        for index, row in enumerate(report_soup.table.find_all('tr')):
            
            # first let's get all the elements.
            cols = row.find_all('td')
            
            # if it's a regular row and not a section or a table header
            if (len(row.find_all('th')) == 0 and len(row.find_all('strong')) == 0): 
                reg_row = [ele.text.strip() for ele in cols]
                statement_data['data'].append(reg_row)
                
            # if it's a regular row and a section but not a table header
            elif (len(row.find_all('th')) == 0 and len(row.find_all('strong')) != 0):
                sec_row = cols[0].text.strip()
                statement_data['sections'].append(sec_row)
                
            # finally if it's not any of those it must be a header
            elif (len(row.find_all('th')) != 0):            
                hed_row = [ele.text.strip() for ele in row.find_all('th')]
                statement_data['headers'].append(hed_row)
                
            else:            
                print('We encountered an error.')

        # append it to the master list.
        statements_data.append(statement_data)

    return statements_data

"""
Convert the Data into a Data Frame, then convert to excel:
"""
def make_balSheet_df(statements_data):
    # Grab the proper components
    bal_header =  statements_data[0]['headers'][0]
    bal_data = statements_data[0]['data']
    #print("bal header:\n",bal_header)
    #print("bal data:\n",bal_data)

    # Put the data in a DataFrame
    balSheet_df = pd.DataFrame(bal_data)
    '''
    # Display
    print('-'*100)
    print('Before Reindexing')
    print('-'*100)
    print(balSheet_df.head())
    '''
    # Define the Index column, rename it, and we need to make sure to drop the old column once we reindex.
    balSheet_df.index = balSheet_df[0]
    balSheet_df.index.name = 'Category'
    balSheet_df = balSheet_df.drop(0, axis = 1)
    '''
    # Display
    print('-'*100)
    print('Before Regex')
    print('-'*100)
    print(balSheet_df.head())
    '''
    # Get rid of the '$', '(', ')', and convert the '' to NaNs.
    balSheet_df = balSheet_df.replace('[\$,)]','', regex=True )\
                         .replace( '[(]','-', regex=True)\
                         .replace( '', 'NaN', regex=True)
    '''
    # Display
    print('-'*100)
    print('Before type conversion')
    print('-'*100)
    print(balSheet_df.head())
    '''
    # everything is a string, so let's convert all the data to a float.
    balSheet_df = balSheet_df.astype(float)

    # Change the column headers
    balSheet_df.columns = bal_header #TODO: there is something wrong with this line.
    #Error: ValueError: Length mismatch: Expected axis has 5 elements, new values have 6 elements

    # Display
    #print('-'*100)
    #print('Final Product')
    #print('-'*100)

    #print("Balance Sheet df: ")
    #print(balSheet_df)
    # convert pandas dataframe into excel worksheet
    wsBal = wb.create_sheet("Balance Sheet")
    for r in dataframe_to_rows(balSheet_df, index=True, header=True):
        wsBal.append(r)
    #return wsBal

def make_income_df(statements_data):
    # Grab the proper components
    income_header =  statements_data[1]['headers'][1]
    income_data = statements_data[1]['data']
    #print("income header:\n",income_header)
    #print("income data:\n",income_data)

    # Put the data in a DataFrame
    income_df = pd.DataFrame(income_data)
    '''
    # Display
    print('-'*100)
    print('Before Reindexing')
    print('-'*100)
    print(income_df.head())
    '''
    # Define the Index column, rename it, and we need to make sure to drop the old column once we reindex.
    income_df.index = income_df[0]
    income_df.index.name = 'Category'
    income_df = income_df.drop(0, axis = 1)
    '''
    # Display
    print('-'*100)
    print('Before Regex')
    print('-'*100)
    print(income_df.head())
    '''
    # Get rid of the '$', '(', ')', and convert the '' to NaNs.
    income_df = income_df.replace('[\$,)]','', regex=True )\
                         .replace( '[(]','-', regex=True)\
                         .replace( '', 'NaN', regex=True)
    '''
    # Display
    print('-'*100)
    print('Before type conversion')
    print('-'*100)
    print(income_df.head())
    '''
    # everything is a string, so let's convert all the data to a float.
    income_df = income_df.astype(float)

    # Change the column headers
    income_df.columns = income_header

    # Display
    #print('-'*100)
    #print('Final Product')
    #print('-'*100)

    #print(income_df)
    # convert pandas dataframe into excel worksheet
    wsIncome = wb.create_sheet("Statement of Net Income (Loss)")
    for r in dataframe_to_rows(income_df, index=True, header=True):
        wsIncome.append(r)
    return wsIncome

def make_cashFlow_df(statements_data):
    # Grab the proper components
    cashFlow_header =  statements_data[2]['headers'][1]
    cashFlow_data = statements_data[2]['data']
    #print("Cash flow header:\n",cashFlow_header)
    #print("Cash flow data:\n",cashFlow_data)
    
    # Put the data in a DataFrame
    cashFlow_df = pd.DataFrame(cashFlow_data)
    '''
    # Display
    print('-'*100)
    print('Before Reindexing')
    print('-'*100)
    print(cashFlow_df.head())
    '''
    # Define the Index column, rename it, and we need to make sure to drop the old column once we reindex.
    cashFlow_df.index = cashFlow_df[0]
    cashFlow_df.index.name = 'Category'
    cashFlow_df = cashFlow_df.drop(0, axis = 1)
    '''
    # Display
    print('-'*100)
    print('Before Regex')
    print('-'*100)
    print(cashFlow_df.head())
    '''
    # Get rid of the '$', '(', ')', and convert the '' to NaNs.
    cashFlow_df = cashFlow_df.replace('[\$,)]','', regex=True )\
                         .replace( '[(]','-', regex=True)\
                         .replace( '', 'NaN', regex=True)
    '''
    # Display
    print('-'*100)
    print('Before type conversion')
    print('-'*100)
    print(cashFlow_df.head())
    '''
    # everything is a string, so let's convert all the data to a float.
    cashFlow_df = cashFlow_df.astype(float)

    # Change the column headers
    cashFlow_df.columns = cashFlow_header

    # Display
    #print('-'*100)
    #print('Final Product')
    #print('-'*100)

    #print(cashFlow_df)
    # convert pandas dataframe into excel worksheet
    wsCashFlow = wb.create_sheet("Statement of Cash Flows")
    for r in dataframe_to_rows(cashFlow_df, index=True, header=True):
        wsCashFlow.append(r)
    return wsCashFlow

def make_equity_df(statements_data):
    # Grab the proper components
    equity_header =  statements_data[3]['headers'][0]
    equity_data = statements_data[3]['data']
    #print("Equity header: ",equity_header)
    #print("Equity_data: ",equity_data)
    
    # Put the data in a DataFrame
    equity_df = pd.DataFrame(equity_data)
    '''
    # Display
    print('-'*100)
    print('Before Reindexing')
    print('-'*100)
    print(equity_df.head())
    '''
    # Define the Index column, rename it, and we need to make sure to drop the old column once we reindex.
    equity_df.index = equity_df[0]
    equity_df.index.name = 'Category'
    equity_df = equity_df.drop(0, axis = 1)
    '''
    # Display
    print('-'*100)
    print('Before Regex')
    print('-'*100)
    print(equity_df.head())
    '''
    # Get rid of the '$', '(', ')', and convert the '' to NaNs.
    equity_df = equity_df.replace('[\$,)]','', regex=True )\
                         .replace( '[(]','-', regex=True)\
                         .replace( '', 'NaN', regex=True)
    '''
    # Display
    print('-'*100)
    print('Before type conversion')
    print('-'*100)
    print(equity_df.head())
    '''
    # everything is a string, so let's convert all the data to a float.
    equity_df = equity_df.astype(float)

    # Change the column headers
    equity_df.columns = equity_header #TODO: there is something wrong with this line.
    #ValueError: Length mismatch: Expected axis has 11 elements, new values have 6 elements

    # Display
    #print('-'*100)
    #print('Final Product')
    #print('-'*100)

    #print(equity_df)
    # convert pandas dataframe into excel worksheet
    wsEquity = wb.create_sheet("Statement of Stockholders' Equity")
    for r in dataframe_to_rows(equity_df, index=True, header=True):
       wsEquity.append(r)
    #return wsEquity

"""
Perform Financial Analyses
"""



"""
Format Excel Document
"""
#def format_balSheet():
    

def format_income(ws):
    ws.insert_rows(1)
    ws.merge_cells('A1:L1') # TODO: have the program automatically
    #merge the top row based on how many columns there are in the given worksheet
    
    ws["A1"] = "Company Income Statement" # TODO: create a variable that automatically
    #prints the company's name rather than the word "Company"

    # adjust column widths
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    # give data accounting format
    #ws["B3:L22"].BUILTIN_FORMATS[8] # TODO: make this cell range variable

    # center the cells that should be centered
    ws["A1"].alignment = Alignment(horizontal='center')
    ws["B2:L22"].alignment = Alignment(horizontal='center')

    # bold the cells that should be bolded
    ws["A1:L2"].bold

def format_cashFlow(ws):
    ws.insert_rows(1)
    ws.merge_cells('A1:D1') # TODO: have the program automatically
    #merge the top row based on how many columns there are in the given worksheet

    ws["A1"] = "Company Statement of Cash Flows" # TODO: create a variable that automatically
    #prints the company's name rather than the word "Company"

    # adjust column widths
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    

#def format_equity():
    

#def format_analysis():
    

"""
Name and Main functions
"""
def main():
    #get_filing_summary()
    xml_summary = get_filing_summary()
    
    #parse_filing_summary()
    master_reports = parse_filing_summary(xml_summary)

    #grab_financial_statements()
    statements_url = grab_financial_statements(master_reports)

    #scrape_financial_statements()
    statements_data = scrape_financial_statements(statements_url)

    '''
    #make_balSheet_df(statements_data)
    #print("\n\n")
    wsIncome = make_income_df(statements_data)
    #print("\n\n")
    wsCashFlow = make_cashFlow_df(statements_data)
    #print("\n\n")
    #make_equity_df(statements_data)'''

    #format_analysis()
    #format_balSheet()
    #format_income(wsIncome)
    #format_cashFlow(wsCashFlow)
    #format_equity()
    
    #wb.save("finPyXLPractice.xlsx")

if __name__ == "__main__":
    main()
